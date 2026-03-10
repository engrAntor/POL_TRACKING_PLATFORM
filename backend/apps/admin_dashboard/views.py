import csv
import io
from datetime import datetime

from django.db.models import Count, Sum, Q
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from apps.accounts.permissions import IsAdmin
from .models import POLItem, Notification
from .serializers import (
    POLItemSerializer, POLItemCreateSerializer,
    NotificationSerializer,
)


# ══════════════════════════════════════════════════════════════════════════════
# OVERVIEW / DASHBOARD STATS
# ══════════════════════════════════════════════════════════════════════════════
class OverviewView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        user = request.user
        today = timezone.now().date()

        total_items = POLItem.objects.filter(user=user).count()
        near_expiry = POLItem.objects.filter(
            user=user, expiry_status='near_expiry'
        ).count()
        low_stock = POLItem.objects.filter(user=user, status='low_stock').count()
        active_transactions = user.orders.filter(
            status__in=['pending', 'approved']
        ).count()

        # Stock levels by status
        stock_levels = (
            POLItem.objects.filter(user=user)
            .values('status')
            .annotate(total_qty=Sum('quantity'))
            .order_by('status')
        )

        # Recent orders
        from apps.superadmin.models import Order
        recent_orders = Order.objects.filter(user=user).select_related('user')[:10]
        orders_data = [{
            'id': o.id,
            'product_name': o.product_name,
            'quantity': str(o.quantity),
            'status': o.status,
            'updated_at': o.updated_at.isoformat(),
        } for o in recent_orders]

        return Response({
            'total_pol_items': total_items,
            'near_expiry': near_expiry,
            'low_stock': low_stock,
            'active_transactions': active_transactions,
            'stock_levels': list(stock_levels),
            'recent_orders': orders_data,
        })


# ══════════════════════════════════════════════════════════════════════════════
# USAGE TRACKER (POL Items - full CRUD)
# ══════════════════════════════════════════════════════════════════════════════
class TrackerListView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = POLItemSerializer
    pagination_class = None
    filterset_fields = ['status', 'expiry_status']
    search_fields = ['product_name', 'part_number']
    ordering_fields = ['expiry', 'created_at', 'product_name']

    def get_queryset(self):
        return POLItem.objects.filter(user=self.request.user)


class TrackerCreateView(generics.CreateAPIView):
    permission_classes = [IsAdmin]
    serializer_class = POLItemCreateSerializer
    parser_classes = [MultiPartParser, FormParser]


class TrackerDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdmin]
    serializer_class = POLItemSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return POLItem.objects.filter(user=self.request.user)


# ══════════════════════════════════════════════════════════════════════════════
# INVENTORY (same POL Items, different filters)
# ══════════════════════════════════════════════════════════════════════════════
class InventoryListView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = POLItemSerializer
    pagination_class = None
    filterset_fields = ['status', 'expiry_status']
    search_fields = ['product_name', 'part_number']

    def get_queryset(self):
        return POLItem.objects.filter(user=self.request.user)


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════
class NotificationListView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = NotificationSerializer
    pagination_class = None

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)


class TriggerExpiryAlertsView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        from .tasks import generate_expiry_alerts
        count = generate_expiry_alerts(request.user)
        return Response({'message': f'{count} new expiry alerts generated.'})


# ══════════════════════════════════════════════════════════════════════════════
# BULK FILE UPLOAD (CSV / XLS / XLSX)
# ══════════════════════════════════════════════════════════════════════════════
class BulkCSVUploadView(APIView):
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    ALLOWED_EXTENSIONS = ('.csv', '.xls', '.xlsx')

    def _parse_rows(self, file):
        """Return a list of dicts from CSV or Excel file."""
        name = file.name.lower()

        if name.endswith('.csv'):
            decoded = file.read().decode('utf-8')
            return list(csv.DictReader(io.StringIO(decoded)))

        if name.endswith(('.xls', '.xlsx')):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else '' for h in next(rows_iter)]
            result = []
            for row in rows_iter:
                result.append({
                    headers[i]: str(cell).strip() if cell is not None else ''
                    for i, cell in enumerate(row) if i < len(headers)
                })
            wb.close()
            return result

        return None

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.lower().endswith(self.ALLOWED_EXTENSIONS):
            return Response(
                {'error': 'Only CSV, XLS, and XLSX files are accepted.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = self._parse_rows(file)
            if rows is None:
                raise ValueError('Unsupported format')
        except Exception:
            return Response(
                {'error': 'Could not read file. Ensure it is a valid CSV or Excel file.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        errors = []
        today = timezone.now().date()

        for i, row in enumerate(rows, start=2):
            try:
                # Required columns
                part_number = row.get('Part Number', row.get('Partnumber', '')).strip()
                description = row.get('Description', row.get('Part Description', '')).strip()
                pol_type = row.get('Type', '').strip().lower() or 'petroleum'
                uom = row.get('UOM', row.get('Unit of Measurement', '')).strip().upper() or 'GAL'
                quantity = row.get('Stock/Availability', row.get('QREQ', '0')).strip()
                shelf_life_raw = row.get('Shelf Life', row.get('SHELF LIFE', '0')).strip()
                good_till = row.get('Expiry', row.get('GOOD TILL', '')).strip()
                condition_raw = row.get('Condition', '').strip().lower().replace(' ', '_') or 'new_pol'
                condition = condition_raw
                price = row.get('Price', '0').strip()

                # Optional columns
                product_name = row.get('Product Name', '').strip()
                alt_part_number = row.get('Alt Part Number', '').strip()
                manufacturer_part_number = row.get('Manufacturer Part Number', '').strip()
                mil_spec = row.get('MIL Spec', '').strip()
                serial_number = row.get('Serial Number', '').strip()
                batch_number = row.get('Batch Number', '').strip()
                source = row.get('Source', '').strip()
                balance = row.get('Balance', row.get('Usage Rate', '')).strip()
                notes = row.get('Notes', '').strip()

                if not part_number:
                    errors.append(f"Row {i}: Missing part number.")
                    continue

                # Parse expiry date (try multiple formats)
                expiry_date = None
                for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        expiry_date = datetime.strptime(good_till, fmt).date()
                        break
                    except ValueError:
                        continue
                if not expiry_date:
                    errors.append(f"Row {i}: Invalid or missing expiry date '{good_till}'.")
                    continue

                # Shelf life string
                shelf_life_str = shelf_life_raw
                if shelf_life_raw.isdigit():
                    shelf_life_str = f"{shelf_life_raw} days"

                # Determine expiry status
                days_left = (expiry_date - today).days
                if days_left < 0:
                    expiry_status = 'expired'
                elif days_left <= 90:
                    expiry_status = 'near_expiry'
                else:
                    expiry_status = 'active'

                # Validate type and condition against choices
                valid_types = ['petroleum', 'oil', 'lubricant']
                if pol_type not in valid_types:
                    pol_type = 'petroleum'
                valid_conditions = ['new_pol', 'leftover_pol', 'opened_pol']
                if condition not in valid_conditions:
                    condition = 'new_pol'
                valid_uoms = ['QT', 'OZ', 'LB', 'RL', 'EA', 'GAL', 'ML', 'PT', 'KT', 'GM', 'FT', 'SQ ST', 'YD', 'CC']
                if uom not in valid_uoms:
                    uom = 'GAL'

                item_data = dict(
                    description=description,
                    pol_type=pol_type,
                    uom=uom,
                    quantity=float(quantity) if quantity else 0,
                    shelf_life=shelf_life_str,
                    expiry_status=expiry_status,
                    condition=condition,
                    price_per_unit=float(price) if price else 0,
                    product_name=product_name,
                    alt_part_number=alt_part_number,
                    manufacturer_part_number=manufacturer_part_number,
                    mil_spec=mil_spec,
                    serial_number=serial_number,
                    batch_number=batch_number,
                    source=source,
                    balance=balance,
                    notes=notes,
                )

                # Override if Part Number + Expiry match, otherwise create new
                existing = POLItem.objects.filter(
                    user=request.user,
                    part_number=part_number,
                    expiry=expiry_date,
                ).first()

                if existing:
                    for key, val in item_data.items():
                        setattr(existing, key, val)
                    existing.save()
                    updated += 1
                else:
                    POLItem.objects.create(
                        user=request.user,
                        part_number=part_number,
                        expiry=expiry_date,
                        **item_data,
                    )
                    created += 1
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")

        return Response({
            'message': f'{created} items imported, {updated} items updated.',
            'created': created,
            'updated': updated,
            'errors': errors,
        })


# ══════════════════════════════════════════════════════════════════════════════
# MSDS FILE UPLOAD (attach PDF to existing POL item)
# ══════════════════════════════════════════════════════════════════════════════
class MSDSUploadView(APIView):
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        try:
            item = POLItem.objects.get(pk=pk, user=request.user)
        except POLItem.DoesNotExist:
            return Response(
                {'error': 'POL item not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        item.msds_file = file
        item.save()
        return Response({
            'message': 'MSDS file uploaded successfully.',
            'msds_file': request.build_absolute_uri(item.msds_file.url),
        })
